import logging
from openpyxl import load_workbook, worksheet, utils
import pandas as pd
from itertools import *
import numpy as np
from datetime import datetime
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.Logger('catch_all')

EMPLEADO = 'empleado'
SEPARADOR= '!'
empleados = {}
nombre_empleados = []
nombres_unicos = set()
conceptos = []
concentrado_conceptos = []
#indices en minusculas
datos=['empleado']

#rangos = ('hojas_plantilla', 'hojas_honorarios', 'hojas_compensaciones')
rangos = {
            'hojas_plantilla': {
                'conceptos':['ispt', 'isr_bono', 'isr_grat_men', 'isr_aguinaldo']
            },
            'hojas_compensaciones': {
                'conceptos':['isr_compen', 'isr_grat_anual_compen']
            },
            'hojas_honorarios':{
                'conceptos':['isr_grat', 'isr_grat_qnal', 'isr_grat_anual']
            }
        }

"""
funcion que concentra los nombres de los empleados
"""
def empleados_unicos(hoja):
    global nombre_empleados
    #encuentra el indice del nombre del empleado
    df = pd.DataFrame(hoja.values)
    #for col in df:
    for name, values in df.iteritems():
        titulo = values[0]
        if (not titulo is None) and (titulo.lower() == EMPLEADO):
            #print(values[1:])
            nombres = filter(None,values[1:])
            nombres = [x.strip().upper() for x in nombres]
            nombre_empleados.extend(nombres)
            break

""" 
funcion que lee los ramgos establecidos en la hoja de nombres
"""
def recorre_rangos(libro):
    #recorre los rangos definidos
    global conceptos
    global concentrado_conceptos
    for k, v in rangos.items():
        address = list(libro.defined_names[k].destinations)
        #cambia los valores globales de los conceptos a buscar
        conceptos=v['conceptos']
        concentrado_conceptos = concentrado_conceptos + conceptos
        #removing the $ from the address
        for sheetname, cellAddress in address:
            cellAddress = cellAddress.replace('$','')
            worksheet = libro[sheetname]
            celdas_hojas = worksheet[cellAddress]
            recorre_rango_nombres(libro,celdas_hojas)

"""
#funcion que recorre las hojas establecidas en el rango
"""
def recorre_rango_nombres(libro, rango_nombres):
    #global nombres_unicos
    for i in range(0,len(rango_nombres)):
            for nombre in rango_nombres[i]:
                print(nombre.value)
                hoja_nomina = libro[nombre.value]
                #paso 1:
                #Determinar inconsistencias en los nombres
                #empleados_unicos(hoja_nomina)
                #paso 2:
                #generar el indice de los titulos para poder mapera los conceptos
                recorre_hoja(nombre.value,hoja_nomina)

    #convierte la lista en un set
    #paso 1.2: muestra los nombres unicos para depurar excel
    #nombres_unicos=set(nombre_empleados)        


"""
    funcion que recorre todas las hojas y genera el concentrado de referencias
"""
def recorre_hoja(nombre_hoja,hoja_nomina):
    global datos
    global empleados    
    # Put the sheet values in `data`
    columnas = [utils.get_column_letter(x) for x in range(1,hoja_nomina.max_column+1)]
    indices = list(range(1,hoja_nomina.max_row+1))
    df = pd.DataFrame(hoja_nomina.values, columns=columnas, index = indices)
    #elimina las filas vacias
    df = df.loc[:, (df != 0).any(axis=0)]
    #inicializa indices
    indices={}
    indice_empleado = ''
    #recorre todas las filas
    for index,row in df.iterrows():
        #recupera los indices en la primer fila
        if index == 1:
            for i,encabezado in enumerate(row, start=1):
                #verifica que el encabezado este dentro de los indices
                if not encabezado is None and isinstance(encabezado, str):
                    encabezado = encabezado.strip().lower()
                    if encabezado in datos + conceptos:
                        indices[encabezado] = utils.get_column_letter(i)
            if not EMPLEADO in indices:
                break
            indice_empleado = indices.pop(EMPLEADO, None)
            #print(indices)
        else:
            #recorre los demas registros del dataset
            nombre = row[indice_empleado]
            if not nombre is None:
                nombre = nombre.strip().upper()
                #se recupera el diccionario de los conceptos por empleado
                if not nombre in empleados:
                    empleado = {
                            'empleado': nombre,
                            'conceptos':{}
                    }
                else:
                    empleado = empleados[nombre]
                #recorre todos los indices
                for k, v in indices.items():
                    valor = row[v]
                    if k in conceptos:
                        if not valor is None and valor > 0:
                            if k in empleado['conceptos']:
                                referencias = empleado['conceptos'][k]
                            else:
                                referencias = []
                            #calcula la direcion de la celda conforme al dataset
                            referencia = "'{0}'{1}{2}{3}".format(nombre_hoja, SEPARADOR, v, index)
                            #agrega la referencia a la lista
                            referencias.append(referencia)
                            empleado['conceptos'][k] = referencias
                    else:
                        empleado[k] = valor
                #actualiza el valor del empleado
                empleados[nombre] = empleado

def genera_df_sumas():
    columnas= list(datos + concentrado_conceptos)
    #inicializa el dataset
    df_empleados = pd.DataFrame(columns=columnas)#, index= [x for x in range(0, len(empleados))])
    #imprime nombre en el dataset
    index = 0
    for nombre, empleado in empleados.items():
        dic_empleado = {}
        #for propiedad, valor in empleado.items():
        for d in datos:
            if d in empleado:
                print(empleado[d])
                dic_empleado[d] = empleado[d]
            else:
                dic_empleado[d] =None
        for c in concentrado_conceptos:
            if c in empleado['conceptos']:
                celdas = empleado['conceptos'][c]
                dic_empleado[c] = '=SUM({0})'.format(','.join(celdas))
            else:
                dic_empleado[c] = None
        df_emp = pd.DataFrame(dic_empleado, index =[0])
        df_empleados= pd.concat([df_empleados,df_emp], ignore_index = True)
    return df_empleados

if __name__ == "__main__":
    print("Abriendo libro..")
    #wb = load_workbook('CONCENTRADO ANUAL PARA DIM 15.xlsm', read_only=True)
    wb = load_workbook('CONCENTRADO ANUAL PARA DIM 15.xlsm', keep_vba=True)
    recorre_rangos(wb)

    #genera el dataframe conforme al diccionario de empleados
    df = genera_df_sumas()
    #crea una hoja con las sumas
    hoy = datetime.today()
    nombre_hoja = '{0}.{1}.{2}_{3}.{4}'.format(hoy.year,hoy.month,hoy.day,hoy.hour,hoy.minute)
    print(df)
    hoja_sumas = wb.create_sheet(title = nombre_hoja)
    for r in dataframe_to_rows(df, index=True, header=True):
        hoja_sumas.append(r)
    #guarda el libro con las sumas
    wb.save('sumas.xlsm')


